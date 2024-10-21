import json
import openpyxl

with open("b-roll_durations.json", "r") as f:
    broll = json.load(f)

with open("interview_durations.json", "r") as f:
    itv = json.load(f)
    
with open("drama_durations.json", "r") as f:
    drama = json.load(f)

with open("unknown_durations.json", "r") as f:
    unknown = json.load(f)

print(broll["Wagenknecht"])

wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "B-Roll"
ws["A2"] = "Project"
ws["B2"] = "Duration"
for i, (key, value) in enumerate(broll.items(), start=3):
    ws[f"A{i}"] = key
    ws[f"B{i}"] = value

ws["D1"] = "Interview"
ws["D2"] = "Project"
ws["E2"] = "Duration"
for i, (key, value) in enumerate(itv.items(), start=3):
    ws[f"D{i}"] = key
    ws[f"E{i}"] = value

ws["G1"] = "Drama"
ws["G2"] = "Project"
ws["H2"] = "Duration"
for i, (key, value) in enumerate(drama.items(), start=3):
    ws[f"G{i}"] = key
    ws[f"H{i}"] = value

ws["J1"] = "Unknown"
ws["J2"] = "Project"
ws["K2"] = "Duration"
for i, (key, value) in enumerate(unknown.items(), start=3):
    ws[f"J{i}"] = key
    ws[f"K{i}"] = value

wb.save("durations.xlsx")